from fastapi import FastAPI, APIRouter, UploadFile, File, Form, HTTPException, Depends, status, Request
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials, OAuth2PasswordBearer, OAuth2PasswordRequestForm
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, date, timedelta
import pandas as pd
import io
import json
import re
from collections import defaultdict
import pdfplumber
import PyPDF2
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from passlib.context import CryptContext
from jose import JWTError, jwt
import secrets
import aiosmtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import random
import string
from authlib.integrations.starlette_client import OAuth
import httpx

ROOT_DIR = Path(__file__).parent
# Load environment variables - prefer .env.local for development
load_dotenv(ROOT_DIR / '.env.local')  # Load local secrets first
load_dotenv(ROOT_DIR / '.env')        # Load template as fallback

# Security configuration
SECRET_KEY = os.environ.get("SECRET_KEY", secrets.token_urlsafe(32))
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30 * 24 * 60  # 30 days

# Gmail SMTP Configuration
GMAIL_EMAIL = os.environ.get("GMAIL_EMAIL")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD")

# Google OAuth Configuration
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET")

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="auth/login")

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Add session middleware for OAuth
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer(auto_error=False)

# Define Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    household_id: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class UserCreate(BaseModel):
    name: str
    email: str
    household_name: Optional[str] = None

class Category(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    color: str = "#3B82F6"
    user_id: str
    household_id: Optional[str] = None
    is_default: bool = False
    created_at: datetime = Field(default_factory=datetime.utcnow)

class CategoryCreate(BaseModel):
    name: str
    color: str = "#3B82F6"

class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    color: Optional[str] = None

class Transaction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: date
    description: str
    category: str
    amount: float
    account_type: str = "credit_card"
    user_id: str
    household_id: Optional[str] = None
    pdf_source: Optional[str] = None  # Track if imported from PDF and source filename
    user_name: Optional[str] = None  # Extracted user name from PDF
    created_at: datetime = Field(default_factory=datetime.utcnow)

class TransactionCreate(BaseModel):
    date: date
    description: str
    category: str
    amount: float
    account_type: str = "credit_card"
    user_id: Optional[str] = "default_user"

class MonthlyReport(BaseModel):
    month: str
    year: int
    categories: dict
    total_spent: float
    transaction_count: int

class CategorySpending(BaseModel):
    category: str
    amount: float
    count: int
    percentage: float

# Authentication Models
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    full_name: str
    username: Optional[str] = None

class User(BaseModel):
    id: str
    email: str
    username: str
    full_name: str
    role: str = "user"
    household_id: Optional[str] = None
    created_at: datetime
    last_login: Optional[datetime] = None

class UserLogin(BaseModel):
    email: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str

class TokenData(BaseModel):
    email: Optional[str] = None

# Household Models
class HouseholdCreate(BaseModel):
    name: str

class Household(BaseModel):
    id: str
    name: str
    created_by: str
    members: List[str]
    created_at: datetime

# Password Reset Models
class PasswordResetRequest(BaseModel):
    email: EmailStr

class PasswordResetConfirm(BaseModel):
    email: EmailStr
    reset_code: str
    new_password: str

# User Profile Models
class UserProfileUpdate(BaseModel):
    full_name: Optional[str] = None
    username: Optional[str] = None

class PasswordChangeRequest(BaseModel):
    current_password: str
    new_password: str

# Authentication utility functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_user_by_email(email: str):
    user = await db.users.find_one({"email": email})
    if user:
        return user
    return None

async def authenticate_user(email: str, password: str):
    user = await get_user_by_email(email)
    if not user:
        return False
    if not verify_password(password, user["password_hash"]):
        return False
    return user

# Email utility functions
async def send_email(to_email: str, subject: str, body: str):
    """Send email using Gmail SMTP"""
    try:
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = GMAIL_EMAIL
        message["To"] = to_email
        
        # Create the HTML content
        html_part = MIMEText(body, "html")
        message.attach(html_part)
        
        # Send the email
        await aiosmtplib.send(
            message,
            hostname="smtp.gmail.com",
            port=587,
            start_tls=True,
            username=GMAIL_EMAIL,
            password=GMAIL_APP_PASSWORD,
        )
        return True
    except Exception as e:
        logging.error(f"Failed to send email: {e}")
        return False

def generate_reset_code():
    """Generate a 6-digit reset code"""
    return ''.join(random.choices(string.digits, k=6))

async def get_current_user(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise credentials_exception
        token_data = TokenData(email=email)
    except JWTError:
        raise credentials_exception
    user = await get_user_by_email(email=token_data.email)
    if user is None:
        raise credentials_exception
    return user

async def get_current_user_id(current_user: dict = Depends(get_current_user)):
    return current_user["id"]

# Enhanced function for multi-user support
async def get_view_user_id(
    view_user_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get the user ID for data viewing - supports switching between household members"""
    if not view_user_id or view_user_id == 'personal':
        return current_user["id"]
    
    if view_user_id == 'family_view':
        # Return special identifier for family view
        return 'family_view'
    
    # Check if the requested user is in the same household
    if current_user.get("household_id"):
        target_user = await db.users.find_one({"id": view_user_id})
        if target_user and target_user.get("household_id") == current_user["household_id"]:
            return view_user_id
    
    # Default to current user if not authorized
    return current_user["id"]

# Migration compatibility function  
async def get_current_user_id_flexible(authorization: Optional[str] = Depends(oauth2_scheme)):
    """Flexible user ID function for migration - supports both auth and legacy"""
    if not authorization:
        return "default_user"  # Legacy support
    
    try:
        current_user = await get_current_user(authorization)
        return current_user["id"]
    except HTTPException:
        return "default_user"  # Fallback for invalid tokens

# PDF Processing Functions
def extract_text_from_pdf(file_content: bytes) -> str:
    """Extract text from PDF using multiple methods for better reliability"""
    text = ""
    
    try:
        # Method 1: Using pdfplumber (better for tables and structured data)
        with pdfplumber.open(io.BytesIO(file_content)) as pdf:
            print(f"PDF has {len(pdf.pages)} pages")
            for page_num, page in enumerate(pdf.pages):
                print(f"Processing page {page_num + 1}")
                page_text = page.extract_text()
                if page_text:
                    print(f"Page {page_num + 1} extracted {len(page_text)} characters")
                    text += f"\n--- PAGE {page_num + 1} ---\n" + page_text + "\n"
                else:
                    print(f"Page {page_num + 1} - no text extracted")
                    
                # Also try to extract tables
                tables = page.extract_tables()
                if tables:
                    print(f"Page {page_num + 1} has {len(tables)} tables")
                    for table_num, table in enumerate(tables):
                        text += f"\n--- TABLE {table_num + 1} ON PAGE {page_num + 1} ---\n"
                        for row in table:
                            if row and any(cell for cell in row if cell):  # Skip empty rows
                                text += " | ".join(str(cell) if cell else "" for cell in row) + "\n"
    
    except Exception as e:
        logging.warning(f"pdfplumber extraction failed: {e}")
        
        # Method 2: Fallback to PyPDF2
        try:
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
            print(f"PyPDF2: PDF has {len(pdf_reader.pages)} pages")
            for page_num, page in enumerate(pdf_reader.pages):
                page_text = page.extract_text()
                text += f"\n--- PYPDF2 PAGE {page_num + 1} ---\n" + page_text + "\n"
        except Exception as e2:
            logging.error(f"PyPDF2 extraction also failed: {e2}")
            raise HTTPException(status_code=400, detail="Could not extract text from PDF")
    
    print(f"Total extracted text length: {len(text)}")
    print("EXTRACTED TEXT PREVIEW:")
    print("=" * 80)
    print(text[:2000])  # Print first 2000 characters for debugging
    print("=" * 80)
    
    return text

def generate_source_name(user_name: str, account_type: str, original_filename: str = None) -> str:
    """Generate a user-friendly source name like 'JANE'S DEBIT' or 'JOHN'S CREDIT'"""
    if not user_name or user_name == 'Unknown User':
        # Fallback to filename if no user name found
        if original_filename:
            base_name = original_filename.replace('.pdf', '').replace('_', ' ').title()
            return f"{base_name} ({account_type.title()})"
        return f"Unknown {account_type.title()}"
    
    # Clean up the user name
    name_parts = user_name.strip().split()
    if len(name_parts) >= 2:
        # Use first name for possessive form
        first_name = name_parts[0].title()
        return f"{first_name}'s {account_type.title()}"
    else:
        # Single name or short name
        clean_name = user_name.title()
        return f"{clean_name}'s {account_type.title()}"

def extract_pdf_metadata(text: str) -> dict:
    """Extract user name and statement period from PDF header - enhanced for multiple formats"""
    metadata = {
        'user_name': None,
        'statement_start': None,
        'statement_end': None,
        'statement_year': None
    }
    
    lines = text.split('\n')
    
    # Enhanced user name extraction for both debit and credit formats
    for i, line in enumerate(lines[:15]):  # Look at more lines
        line = line.strip()
        
        # Method 1: Look for name patterns (all caps, likely a person's name)
        if re.match(r'^[A-Z\s]{8,50}$', line) and ' ' in line and len(line.split()) >= 2:
            # Skip common bank terms
            bank_terms = ['ACCOUNT', 'STATEMENT', 'CARD', 'BANK', 'DIVIDEND', 'CIBC', 'VISA', 'TRANSACTION', 'DETAILS']
            if not any(term in line.upper() for term in bank_terms):
                metadata['user_name'] = line.strip()
                print(f"Found user name (Method 1): {metadata['user_name']}")
                break
        
        # Method 2: Look for "Prepared for:" pattern (credit cards)
        prepared_match = re.search(r'prepared for:?\s*([A-Z\s]+?)(?:\s+[A-Z]{2,}\s+\d|\s*$)', line, re.IGNORECASE)
        if prepared_match:
            name = prepared_match.group(1).strip()
            if len(name) > 5 and ' ' in name:  # Valid name should have space and be reasonable length
                metadata['user_name'] = name
                print(f"Found user name (Method 2): {metadata['user_name']}")
                break
        
        # Method 3: Look for name right after date pattern (debit format)
        # Example: "JANE AGBAOHWO                                For Jul 1 to Jul 31, 2024"
        date_with_name = re.search(r'^([A-Z\s]+?)\s+For\s+(\w+\s+\d+\s+to\s+\w+\s+\d+,?\s+\d{4})', line, re.IGNORECASE)
        if date_with_name:
            name = date_with_name.group(1).strip()
            if len(name) > 5 and ' ' in name:
                metadata['user_name'] = name
                print(f"Found user name (Method 3): {metadata['user_name']}")
                # Also extract the date from this line
                date_part = date_with_name.group(2)
                period_match = re.search(r'(\w+)\s+(\d+)\s*to\s*(\w+)\s+(\d+),?\s*(\d{4})', date_part, re.IGNORECASE)
                if period_match:
                    start_month, start_day, end_month, end_day, year = period_match.groups()
                    metadata['statement_start'] = f"{start_month} {start_day}"
                    metadata['statement_end'] = f"{end_month} {end_day}"
                    metadata['statement_year'] = int(year)
                break
    
    # Look for statement period if not found above
    if not metadata['statement_year']:
        for line in lines[:20]:
            line = line.strip()
            # Pattern: "October 16to November 15, 2024" or "October 16 to November 15, 2024"
            period_match = re.search(r'(\w+)\s+(\d+)\s*to\s*(\w+)\s+(\d+),?\s*(\d{4})', line, re.IGNORECASE)
            if period_match:
                start_month, start_day, end_month, end_day, year = period_match.groups()
                metadata['statement_start'] = f"{start_month} {start_day}"
                metadata['statement_end'] = f"{end_month} {end_day}"
                metadata['statement_year'] = int(year)
                break
            
            # Alternative pattern: "November 15, 2024" for statement date
            date_match = re.search(r'(\w+)\s+(\d+),?\s*(\d{4})', line)
            if date_match and 'statement' in line.lower():
                month, day, year = date_match.groups()
                metadata['statement_end'] = f"{month} {day}"
                metadata['statement_year'] = int(year)
    
    return metadata

def parse_transactions_from_text(text: str, user_id: str, source_filename: str = None) -> List[dict]:
    """Parse transactions from extracted PDF text - enhanced for multiple bank formats"""
    transactions = []
    
    # Extract metadata first
    metadata = extract_pdf_metadata(text)
    statement_year = metadata.get('statement_year', datetime.now().year)
    user_name = metadata.get('user_name', 'Unknown User')
    
    print(f"Extracted metadata: User: {user_name}, Year: {statement_year}")
    
    # Enhanced patterns for CIBC format specifically
    # The correct format should be: Trans Date, Post Date, Description, Category, Amount
    patterns = [
        # Main CIBC pattern - Trans Date, Post Date, Description, Spend Category, Amount
        r'(\w{3}\s+\d{1,2})\s+(\w{3}\s+\d{1,2})\s+([A-Z0-9\s\#\*\.\-\/\&\(\)]+?)\s+([A-Za-z\s,&]+?)\s+(\d+\.\d{2})(?:\s|$)',
        # Alternative pattern for transactions without clear category
        r'(\w{3}\s+\d{1,2})\s+(\w{3}\s+\d{1,2})\s+([A-Z0-9\s\#\*\.\-\/\&\(\)]+?)\s+(\d+\.\d{2})(?:\s|$)',
        # Catch remaining patterns
        r'(\w{3}\s+\d{1,2})\s+([A-Z0-9\s\#\*\.\-\/\&\(\)]{10,})\s+(\d+\.\d{2})(?:\s|$)'
    ]
    
def detect_statement_format(text: str) -> str:
    """Detect the type of CIBC statement format"""
    text_upper = text.upper()
    
    # Check for debit account indicators
    debit_indicators = [
        'TRANSACTION DETAILS',
        'WITHDRAWALS ($)',
        'DEPOSITS ($)',
        'BALANCE ($)',
        'VISA DEBIT RETAIL PURCHASE',
        'ACCOUNT SUMMARY'
    ]
    
    # Check for credit card indicators  
    credit_indicators = [
        'YOUR NEW CHARGES',
        'SPEND CATEGORIES',
        'CARD NUMBER',
        'DIVIDEND',
        'VISA CARD'
    ]
    
    debit_score = sum(1 for indicator in debit_indicators if indicator in text_upper)
    credit_score = sum(1 for indicator in credit_indicators if indicator in text_upper)
    
    print(f"Format detection - Debit score: {debit_score}, Credit score: {credit_score}")
    
    if debit_score > credit_score:
        return 'debit'
    else:
        return 'credit'

def parse_cibc_debit_transactions(text: str, user_id: str, source_filename: str, statement_year: int, user_name: str) -> List[dict]:
    """Parse CIBC debit account transactions with table format"""
    transactions = []
    
    print("Parsing CIBC debit format...")
    
    lines = text.split('\n')
    in_transaction_section = False
    
    for line_num, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue
            
        # Look for transaction details section
        if 'transaction details' in line.lower():
            in_transaction_section = True
            print(f"Found transaction details section at line {line_num}")
            continue
            
        # Skip header lines
        if any(header in line.upper() for header in ['DATE', 'DESCRIPTION', 'WITHDRAWALS', 'DEPOSITS', 'BALANCE']):
            continue
            
        # Stop at end indicators
        if any(end_marker in line.lower() for end_marker in ['closing balance', 'important:', 'free transaction']):
            break
            
        if in_transaction_section and line:
            # Try to parse debit transaction line
            # Format: Date | Description | Withdrawals | Deposits | Balance
            
            # Look for date pattern at start of line
            date_match = re.match(r'^(\w{3}\s+\d{1,2})', line)
            if not date_match:
                continue
                
            date_str = date_match.group(1)
            remaining_line = line[len(date_str):].strip()
            
            # Look for amounts (withdrawals, deposits, balance)
            # Balance is usually at the end
            balance_match = re.search(r'(\d+\.\d{2})\s*$', remaining_line)
            if not balance_match:
                continue
                
            balance_amount = balance_match.group(1)
            line_without_balance = remaining_line[:balance_match.start()].strip()
            
            # Look for withdrawal or deposit amount before balance
            amount_pattern = r'(\d+\.\d{2})\s+(\d+\.\d{2})?'
            amounts = re.findall(r'(\d+\.\d{2})', line_without_balance)
            
            if not amounts:
                continue
                
            # The transaction amount is typically the first amount found
            # Enhanced to handle negative amounts and credits
            amount_str = amounts[0].strip()
            
            # Check for negative sign (indicates credit/deposit)
            is_credit = False
            if amount_str.startswith('-') or amount_str.startswith('('):
                is_credit = True
                amount_str = amount_str.replace('-', '').replace('(', '').replace(')', '').strip()
            
            transaction_amount = float(amount_str)
            
            # For debit accounts: negative usually means deposit/credit, positive means withdrawal/debit
            if is_credit:
                transaction_amount = -transaction_amount
                print(f"💳 DEBIT CREDIT/DEPOSIT detected: ${abs(transaction_amount)} (stored as negative)")
            else:
                print(f"💰 DEBIT WITHDRAWAL detected: ${transaction_amount} (stored as positive)")
            
            # Extract description (everything between date and amounts)
            desc_end_pos = line_without_balance.rfind(amounts[0])
            if desc_end_pos == -1:
                continue
                
            description = line_without_balance[:desc_end_pos].strip()
            
            # Clean up description
            description = re.sub(r'\s+', ' ', description)
            
            # Skip if description is too short or looks like header
            if len(description) < 5:
                continue
                
            # Skip certain transaction types
            skip_keywords = ['balance forward', 'opening balance', 'service charge']
            if any(keyword in description.lower() for keyword in skip_keywords):
                continue
                
            # Parse date
            transaction_date = parse_date_string(date_str, statement_year)
            if not transaction_date:
                continue
                
            # Categorize transaction
            category = clean_category("", description)
            
            # Create enhanced source name
            enhanced_source = generate_source_name(user_name, 'debit', source_filename)
            
            # Create transaction
            transaction = {
                'date': transaction_date.isoformat(),
                'description': description,
                'category': category,
                'amount': transaction_amount,
                'account_type': 'debit',
                'user_id': user_id,
                'pdf_source': enhanced_source,
                'user_name': user_name
            }
            
            transactions.append(transaction)
            print(f"✅ DEBIT ADDED: {description} -> ${transaction_amount} on {transaction_date} ({category})")
    
    return transactions

def parse_transactions_from_text(text: str, user_id: str, source_filename: str = None) -> List[dict]:
    """Parse transactions from extracted PDF text - enhanced for multiple CIBC formats"""
    transactions = []
    
    # Extract metadata first
    metadata = extract_pdf_metadata(text)
    statement_year = metadata.get('statement_year', datetime.now().year)
    user_name = metadata.get('user_name', 'Unknown User')
    
    print(f"Extracted metadata: User: {user_name}, Year: {statement_year}")
    
    # Detect statement format
    format_type = detect_statement_format(text)
    print(f"Detected format: {format_type}")
    
    if format_type == 'debit':
        # Use debit parsing logic
        transactions = parse_cibc_debit_transactions(text, user_id, source_filename, statement_year, user_name)
        print(f"\n=== DEBIT PARSING COMPLETE: {len(transactions)} transactions found ===")
        return remove_duplicates(transactions)
    
    # Original credit card parsing logic
    sections = text.split('--- PAGE')
    
    for section_num, section in enumerate(sections):
        if not section.strip():
            continue
            
        print(f"\n=== PROCESSING SECTION {section_num} ===")
        print(f"Section preview: {section[:500]}...")
        
        lines = section.split('\n')
        
        for line_num, line in enumerate(lines):
            line = line.strip()
            if not line or len(line) < 15:
                continue
                
            # Debug: Print every non-empty line to catch missing transactions
            print(f"LINE {line_num}: {line}")
                
            # Skip header lines and section headers - be more specific
            skip_line = False
            line_upper = line.upper()
            
            # Only skip if the line is clearly a header (contains multiple header keywords or exact matches)
            # Use word boundaries to avoid false positives (e.g., "LOVISA" containing "VISA")
            header_keywords = [r'\bCARD NUMBER\b', r'\bPAGE\b', r'\bCIBC\b', r'\bDIVIDEND\b', r'\bVISA\b', r'\bYOUR PAYMENTS\b', r'\bYOUR NEW CHARGES\b']
            table_headers = ['TRANS   POST', 'DATE    DATE', 'SPEND CATEGORIES', 'AMOUNT($)']
            
            # Skip if it's clearly a header line
            skip_reason = None
            if any(re.search(header, line_upper) for header in header_keywords):
                skip_line = True
                matching_headers = [h for h in header_keywords if re.search(h, line_upper)]
                skip_reason = f"Contains header keyword: {matching_headers}"
            elif any(header in line_upper for header in table_headers):
                skip_line = True
                skip_reason = f"Contains table header: {[h for h in table_headers if h in line_upper]}"
            elif line_upper.strip() in ['TRANS', 'POST', 'DESCRIPTION', 'AMOUNT', 'SPEND CATEGORIES']:
                skip_line = True
                skip_reason = f"Exact match header: {line_upper.strip()}"
            
            # Special debug for Lovisa line
            if 'lovisa' in line.lower():
                print(f"🔍 LOVISA DEBUG: skip_line={skip_line}, skip_reason={skip_reason}")
                print(f"🔍 LOVISA LINE_UPPER: '{line_upper}'")
                print(f"🔍 LOVISA HEADER_KEYWORDS: {[h for h in header_keywords if h in line_upper]}")
                print(f"🔍 LOVISA TABLE_HEADERS: {[h for h in table_headers if h in line_upper]}")
            
            if skip_line:
                print(f"SKIPPED HEADER: {line} (Reason: {skip_reason})")
                continue
            
            # More aggressive transaction detection
            # Any line with a month abbreviation followed by digits AND a decimal amount
            has_month_day = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2}', line)
            has_decimal_amount = re.search(r'\d+\.\d{2}', line)
            
            # Additional check: if line contains "Lovisa" specifically, flag it for debugging
            if 'lovisa' in line.lower():
                print(f"🔍 FOUND LOVISA LINE: {line}")
                print(f"   Has month/day: {bool(has_month_day)}")
                print(f"   Has decimal amount: {bool(has_decimal_amount)}")
            
            transaction_match = None  # Initialize here
            
            if has_month_day and has_decimal_amount:
                print(f"FOUND TRANSACTION LINE {line_num}: {line}")
                
                transaction_match = None
                
                # Strategy 1: Amount at the end approach (most reliable)
                amount_matches = re.findall(r'(\d{1,3}(?:,\d{3})*\.\d{2})', line)
                if amount_matches:
                    # Take the last amount as the transaction amount
                    amount_str = amount_matches[-1]
                    
                    # Remove commas from amount before converting to float
                    clean_amount_str = amount_str.replace(',', '')
                    
                    # Find where this amount starts in the line
                    amount_pos = line.rfind(amount_str)
                    line_without_amount = line[:amount_pos].strip()
                    
                    # Extract dates from the beginning
                    date_pattern = re.match(r'(\w{3}\s+\d{1,2})\s+(\w{3}\s+\d{1,2})\s+(.+)', line_without_amount)
                    if date_pattern:
                        trans_date_str, post_date_str, description_and_category = date_pattern.groups()
                        
                        # Clean up the description and category part
                        desc_and_cat = description_and_category.strip()
                        
                        # Skip payment transactions - these are not expenses
                        if any(payment_keyword in desc_and_cat.upper() for payment_keyword in [
                            'PAYMENT THANK YOU', 'PAIEMENT MERCI', 'PAYMENT - THANK YOU', 
                            'THANK YOU FOR YOUR PAYMENT', 'PAYMENT RECEIVED'
                        ]):
                            print(f"SKIPPED PAYMENT: {desc_and_cat}")
                            continue
                        
                        # Try to identify where description ends and category begins
                        category_str = ""
                        description = desc_and_cat
                        
                        # Known categories (longest first to avoid partial matches)
                        known_categories = [
                            'Foreign Currency Transactions',
                            'Hotel, Entertainment and Recreation', 
                            'Professional and Financial Services',
                            'Home and Office Improvement',
                            'Personal and Household Expenses',
                            'Health and Education',
                            'Retail and Grocery',
                            'Transportation',
                            'Restaurants'
                        ]
                        
                        # Try to find category at the end
                        for cat in known_categories:
                            if desc_and_cat.endswith(cat):
                                category_str = cat
                                description = desc_and_cat[:-len(cat)].strip()
                                break
                        
                        # If no category found, try to extract from spacing patterns
                        if not category_str:
                            # Look for multiple spaces that might separate description from category
                            parts = re.split(r'\s{2,}', desc_and_cat)
                            if len(parts) >= 2:
                                description = parts[0].strip()
                                category_str = ' '.join(parts[1:]).strip()
                        
                        transaction_match = (trans_date_str, post_date_str, description, category_str, clean_amount_str)
                        print(f"EXTRACTED: Date={trans_date_str}, Desc='{description}', Cat='{category_str}', Amt={clean_amount_str} (original: {amount_str})")
                    else:
                        print(f"Could not extract dates from: {line_without_amount}")
                
                # If we found a transaction match, process it
            
            # If we found a transaction match, process it
            # If we found a transaction match, process it
            if transaction_match:
                try:
                    # Handle both tuple and regex match objects
                    if isinstance(transaction_match, tuple):
                        trans_date_str, post_date_str, description, category_str, amount_str = transaction_match
                    else:
                        trans_date_str, post_date_str, description, category_str, amount_str = transaction_match.groups()
                    
                    print(f"PROCESSING: Trans={trans_date_str}, Post={post_date_str}, Desc='{description}', Cat='{category_str}', Amt={amount_str}")
                except Exception as e:
                    print(f"Error processing transaction match: {e}")
                    continue
                
                try:
                    # Parse amount - Enhanced to handle negative amounts and credits
                    amount_str_clean = amount_str.strip()
                    
                    # Check for negative sign (indicates credit/payment)
                    is_credit = False
                    if amount_str_clean.startswith('-') or amount_str_clean.startswith('('):
                        is_credit = True
                        amount_str_clean = amount_str_clean.replace('-', '').replace('(', '').replace(')', '').strip()
                    
                    amount = float(amount_str_clean)
                    
                    # Apply negative for credits (payments, refunds)
                    if is_credit:
                        amount = -amount
                        print(f"💳 CREDIT/PAYMENT detected: ${abs(amount)} (stored as negative)")
                    else:
                        print(f"💰 DEBIT/CHARGE detected: ${amount} (stored as positive)")
                    
                    if abs(amount) < 0.01 or abs(amount) > 50000:
                        print(f"Skipping amount {amount} (out of range)")
                        continue
                    
                    # Clean description
                    description = re.sub(r'\s+', ' ', description.strip())
                    
                    # Parse transaction date (use trans_date, not post_date!)
                    transaction_date = parse_date_string(trans_date_str, statement_year)
                    if not transaction_date:
                        print(f"Failed to parse date: {trans_date_str}")
                        continue
                    
                    print(f"🗓️ DATE DEBUG: Input='{trans_date_str}' -> Parsed={transaction_date} -> Final ISO={transaction_date.isoformat()}")
                    
                    # Verify we're using transaction date, not post date
                    post_date_parsed = parse_date_string(post_date_str, statement_year)
                    if post_date_parsed:
                        print(f"🔍 COMPARISON: Transaction={transaction_date} vs Post={post_date_parsed} (using Transaction date)")
                    
                    # Clean and categorize
                    category = clean_category(category_str, description)
                    
                    # Create enhanced source name
                    enhanced_source = generate_source_name(user_name, 'credit', source_filename)
                    
                    # Create transaction
                    transaction = {
                        'date': transaction_date.isoformat(),
                        'description': description,
                        'category': category,
                        'amount': amount,
                        'account_type': 'credit_card',
                        'user_id': user_id,
                        'pdf_source': enhanced_source,
                        'user_name': user_name
                    }
                    
                    transactions.append(transaction)
                    print(f"✅ ADDED: {description} -> ${amount} on {transaction_date} ({category})")
                    
                except Exception as e:
                    print(f"❌ Error processing transaction: {e}")
                    continue
            
            # Alternative pattern for table rows with | separators
            elif '|' in line and re.search(r'\d+\.\d{2}', line):
                print(f"TABLE ROW: {line}")
                parts = [p.strip() for p in line.split('|')]
                if len(parts) >= 5:
                    try:
                        # Assume format: trans_date | post_date | description | category | amount
                        trans_date_str = parts[0]
                        description = parts[2] if len(parts) > 2 else ""
                        category_str = parts[3] if len(parts) > 3 else ""
                        amount_str = parts[-1]  # Last part should be amount
                        
                        # Extract numeric amount
                        amount_match = re.search(r'(\d+\.\d{2})', amount_str)
                        if amount_match:
                            amount = float(amount_match.group(1))
                            
                            transaction_date = parse_date_string(trans_date_str, statement_year)
                            if transaction_date and amount > 0.01:
                                # Create enhanced source name
                                enhanced_source = generate_source_name(user_name, 'credit', source_filename)
                                
                                transaction = {
                                    'date': transaction_date.isoformat(),
                                    'description': description.strip(),
                                    'category': category,
                                    'amount': amount,
                                    'account_type': 'credit_card',
                                    'user_id': user_id,
                                    'pdf_source': enhanced_source,
                                    'user_name': user_name
                                }
                                
                                transactions.append(transaction)
                                print(f"✅ TABLE ADDED: {description} -> ${amount} on {transaction_date}")
                                
                    except Exception as e:
                        print(f"❌ Error processing table row: {e}")
                        continue
    
    print(f"\n=== TOTAL TRANSACTIONS FOUND: {len(transactions)} ===")
    
    # Remove duplicates
    unique_transactions = remove_duplicates(transactions)
    
    return unique_transactions

def parse_date_string(date_str: str, statement_year: int) -> date:
    """Parse date string like 'Oct 22' with given year"""
    try:
        month_map = {
            'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
            'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
        }
        
        parts = date_str.strip().split()
        if len(parts) == 2:
            month_name, day = parts
            month = month_map.get(month_name[:3], None)
            if month:
                # If no statement year provided, use current year or 2024 for reasonable defaults
                year = statement_year if statement_year else 2024
                
                # Smart year logic: if we're in July 2025 and see Oct/Nov dates, they're likely from 2024
                current_year = datetime.now().year
                if not statement_year:
                    # For Oct/Nov/Dec dates when we're in 2025, assume they're from 2024 
                    if month in [10, 11, 12]:  # Oct/Nov/Dec
                        year = 2024  # Most bank statements with these months are from 2024
                    elif month in [1, 2, 3] and datetime.now().month > 6:  # Jan/Feb/Mar but we're late in year  
                        year = current_year + 1
                    else:
                        year = current_year
                
                # Create date object - this should be the EXACT transaction date
                parsed_date = date(year, month, int(day))
                print(f"Date parsing: '{date_str}' -> {parsed_date} (year context: {year})")
                return parsed_date
    except Exception as e:
        print(f"Date parsing error for '{date_str}': {e}")
    return None

def clean_category(category_str: str, description: str) -> str:
    """Clean and standardize category"""
    # Category mapping
    category_keywords = {
        'Retail and Grocery': ['superstore', 'grocery', 'dollarama', 'walmart', 'costco', 'loblaws', 'metro', 'sobeys', 'john & ross', 't&t'],
        'Restaurants': ['restaurant', 'coffee', 'starbucks', 'tim hortons', 'mcdonalds', 'pizza', 'food', 'dining', 'cafe', 'a&w', 'forest lawn'],
        'Transportation': ['lyft', 'uber', 'taxi', 'gas', 'petro', 'shell', 'esso', 'transit', 'ride'],
        'Home and Office Improvement': ['home depot', 'lowes', 'staples', 'canadian tire', 'ikea', 'office', 'stokes'],
        'Hotel, Entertainment and Recreation': ['hotel', 'movie', 'netflix', 'spotify', 'apple.com', 'entertainment', 'apple.com/bill'],
        'Professional and Financial Services': ['bank', 'fee', 'transfer', 'mortgage', 'insurance', 'legal', 'openai', 'chatgpt'],
        'Health and Education': ['pharmacy', 'doctor', 'dental', 'hospital', 'school', 'university'],
        'Foreign Currency Transactions': ['foreign', 'currency', 'exchange', 'international', 'usd']
    }
    
    # First try to use provided category if it looks valid
    if category_str and len(category_str.strip()) > 2:
        category_clean = category_str.strip()
        # Check if it matches known categories
        for known_cat in category_keywords.keys():
            if known_cat.lower() in category_clean.lower():
                return known_cat
    
    # Auto-categorize based on description
    description_lower = description.lower()
    
    for cat, keywords in category_keywords.items():
        if any(keyword in description_lower for keyword in keywords):
            return cat
    
    return 'Personal and Household Expenses'  # Default

def remove_duplicates(transactions: List[dict]) -> List[dict]:
    """Remove duplicate transactions"""
    seen = set()
    unique_transactions = []
    
    for transaction in transactions:
        # Create key using date, first few words of description, and amount
        desc_key = ' '.join(transaction['description'].split()[:3])
        key = (transaction['date'], desc_key, transaction['amount'])
        
        if key not in seen:
            seen.add(key)
            unique_transactions.append(transaction)
        else:
            print(f"Skipping duplicate: {transaction['description']} on {transaction['date']}")
    
    print(f"Total parsed: {len(transactions)}, Unique: {len(unique_transactions)}")
    return unique_transactions

# Initialize default categories
DEFAULT_CATEGORIES = [
    "Retail and Grocery",
    "Restaurants", 
    "Transportation",
    "Home and Office Improvement",
    "Hotel, Entertainment and Recreation",
    "Professional and Financial Services",
    "Health and Education",
    "Foreign Currency Transactions",
    "Personal and Household Expenses"
]

async def initialize_default_categories(user_id: str):
    """Initialize default categories for a user"""
    existing_categories = await db.categories.find({"user_id": user_id}).to_list(100)
    if not existing_categories:
        default_categories = []
        colors = ["#3B82F6", "#10B981", "#F59E0B", "#EF4444", "#8B5CF6", "#06B6D4", "#84CC16", "#F97316", "#6B7280"]
        
        for i, cat_name in enumerate(DEFAULT_CATEGORIES):
            category = Category(
                name=cat_name,
                color=colors[i % len(colors)],
                user_id=user_id,
                is_default=True
            )
            default_categories.append(category.dict())
        
        if default_categories:
            await db.categories.insert_many(default_categories)

# API Routes
@api_router.get("/")
async def root():
    return {"message": "LifeTracker Banking Dashboard API v2.0"}

# User Management
@api_router.post("/users", response_model=User)
async def create_user(user: UserCreate):
    user_data = user.dict()
    user_obj = User(**user_data)
    
    # Convert datetime to string for MongoDB
    user_dict = user_obj.dict()
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    
    await db.users.insert_one(user_dict)
    
    # Initialize default categories for the user
    await initialize_default_categories(user_obj.id)
    
    return user_obj

@api_router.get("/users", response_model=List[User])
async def get_users():
    users = await db.users.find().to_list(100)
    return [User(**user) for user in users]

# Category Management
@api_router.get("/categories", response_model=List[Category])
async def get_categories(user_id: str = Depends(get_current_user_id)):
    await initialize_default_categories(user_id)  # Ensure categories exist
    categories = await db.categories.find({"user_id": user_id}).to_list(100)
    return [Category(**category) for category in categories]

@api_router.post("/categories", response_model=Category)
async def create_category(category: CategoryCreate, user_id: str = Depends(get_current_user_id)):
    category_data = category.dict()
    category_obj = Category(**category_data, user_id=user_id)
    
    # Convert datetime to string for MongoDB
    category_dict = category_obj.dict()
    category_dict['created_at'] = category_dict['created_at'].isoformat()
    
    await db.categories.insert_one(category_dict)
    return category_obj

@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(category_id: str, category_update: CategoryUpdate, user_id: str = Depends(get_current_user_id)):
    update_data = {k: v for k, v in category_update.dict().items() if v is not None}
    
    result = await db.categories.update_one(
        {"id": category_id, "user_id": user_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    
    updated_category = await db.categories.find_one({"id": category_id, "user_id": user_id})
    return Category(**updated_category)

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, user_id: str = Depends(get_current_user_id)):
    # Check if category is being used by transactions
    transactions_using_category = await db.transactions.count_documents({"category": category_id, "user_id": user_id})
    
    if transactions_using_category > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete category. It is used by {transactions_using_category} transactions."
        )
    
    result = await db.categories.delete_one({"id": category_id, "user_id": user_id, "is_default": {"$ne": True}})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found or cannot delete default category")
    
    return {"message": "Category deleted successfully"}

# Transaction Management (Enhanced)
@api_router.post("/transactions", response_model=Transaction)
async def create_transaction(transaction: TransactionCreate, user_id: str = Depends(get_current_user_id)):
    transaction_dict = transaction.dict()
    transaction_dict['user_id'] = user_id
    
    transaction_obj = Transaction(**transaction_dict)
    
    # Convert date to string for MongoDB storage
    transaction_data = transaction_obj.dict()
    transaction_data['date'] = transaction_data['date'].isoformat()
    transaction_data['created_at'] = transaction_data['created_at'].isoformat()
    
    await db.transactions.insert_one(transaction_data)
    return transaction_obj

@api_router.get("/transactions", response_model=List[Transaction])
async def get_transactions(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category: Optional[str] = None,
    pdf_source: Optional[str] = None,
    account_type: Optional[str] = None,
    sort_by: Optional[str] = "date",
    sort_order: Optional[str] = "desc",
    user_id: str = Depends(get_current_user_id)
):
    filter_dict = {"user_id": user_id}
    
    if start_date:
        filter_dict["date"] = {"$gte": start_date}
    if end_date:
        if "date" in filter_dict:
            filter_dict["date"]["$lte"] = end_date
        else:
            filter_dict["date"] = {"$lte": end_date}
    if category:
        filter_dict["category"] = category
    if pdf_source:
        filter_dict["pdf_source"] = pdf_source
    if account_type:
        filter_dict["account_type"] = account_type
    
    # Handle sorting
    sort_direction = -1 if sort_order == "desc" else 1
    sort_field = sort_by if sort_by in ["date", "amount", "description", "category"] else "date"
    
    transactions = await db.transactions.find(filter_dict).sort(sort_field, sort_direction).to_list(1000)
    return [Transaction(**transaction) for transaction in transactions]

@api_router.get("/transactions/sources")
async def get_pdf_sources(user_id: str = Depends(get_current_user_id)):
    """Get list of unique PDF sources for filtering"""
    sources = await db.transactions.distinct("pdf_source", {"user_id": user_id, "pdf_source": {"$ne": None}})
    return {"sources": sources}

@api_router.delete("/transactions/{transaction_id}")
async def delete_transaction(transaction_id: str, user_id: str = Depends(get_current_user_id)):
    result = await db.transactions.delete_one({"id": transaction_id, "user_id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return {"message": "Transaction deleted successfully"}

class TransactionUpdate(BaseModel):
    category: Optional[str] = None
    description: Optional[str] = None
    amount: Optional[float] = None

@api_router.put("/transactions/{transaction_id}")
async def update_transaction(
    transaction_id: str, 
    transaction_update: TransactionUpdate,
    user_id: str = Depends(get_current_user_id)
):
    """Update a transaction's category, description, or amount"""
    update_data = {k: v for k, v in transaction_update.dict().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    result = await db.transactions.update_one(
        {"id": transaction_id, "user_id": user_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Return updated transaction
    updated_transaction = await db.transactions.find_one({"id": transaction_id, "user_id": user_id})
    if updated_transaction:
        # Remove MongoDB's _id field and convert datetime if needed
        if '_id' in updated_transaction:
            del updated_transaction['_id']
        if 'created_at' in updated_transaction and hasattr(updated_transaction['created_at'], 'isoformat'):
            updated_transaction['created_at'] = updated_transaction['created_at'].isoformat()
        return updated_transaction
    else:
        raise HTTPException(status_code=404, detail="Updated transaction not found")

class BulkDeleteRequest(BaseModel):
    transaction_ids: List[str]

@api_router.post("/transactions/bulk-delete")
async def bulk_delete_transactions(
    request: BulkDeleteRequest,
    user_id: str = Depends(get_current_user_id)
):
    """Delete multiple transactions at once"""
    if not request.transaction_ids:
        raise HTTPException(status_code=400, detail="No transaction IDs provided")
    
    result = await db.transactions.delete_many({
        "id": {"$in": request.transaction_ids}, 
        "user_id": user_id
    })
    
    return {
        "message": f"Successfully deleted {result.deleted_count} transactions",
        "deleted_count": result.deleted_count
    }

# Excel Export Endpoint
@api_router.get("/transactions/export/excel")
async def export_transactions_to_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category: Optional[str] = None,
    pdf_source: Optional[str] = None,
    account_type: Optional[str] = None,
    user_id: str = Depends(get_current_user_id)
):
    """Export transactions to Excel file with filters"""
    # Use the same filtering logic as get_transactions
    filter_dict = {"user_id": user_id}
    
    if start_date:
        filter_dict["date"] = {"$gte": start_date}
    if end_date:
        if "date" in filter_dict:
            filter_dict["date"]["$lte"] = end_date
        else:
            filter_dict["date"] = {"$lte": end_date}
    if category:
        filter_dict["category"] = category
    if pdf_source:
        filter_dict["pdf_source"] = pdf_source
    if account_type:
        filter_dict["account_type"] = account_type
    
    # Get transactions sorted by date (newest first)
    transactions = await db.transactions.find(filter_dict).sort("date", -1).to_list(10000)
    
    if not transactions:
        raise HTTPException(status_code=404, detail="No transactions found for export")
    
    # Create Excel workbook
    wb = Workbook()
    
    # Main transactions sheet
    ws_main = wb.active
    ws_main.title = "Transactions"
    
    # Convert transactions to DataFrame for easier manipulation
    df_data = []
    for transaction in transactions:
        df_data.append({
            'Date': transaction.get('date', ''),
            'Description': transaction.get('description', ''),
            'Category': transaction.get('category', ''),
            'Amount': transaction.get('amount', 0),
            'Account Type': 'Credit Card' if transaction.get('account_type') == 'credit_card' else 'Debit Account',
            'Source': transaction.get('pdf_source', 'Manual'),
            'User': transaction.get('user_name', '')
        })
    
    df = pd.DataFrame(df_data)
    
    # Add headers with styling
    headers = ['Date', 'Description', 'Category', 'Amount', 'Account Type', 'Source', 'User']
    for col_num, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for row_num, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_main.cell(row=row_num, column=col_num, value=value)
            if col_num == 4:  # Amount column
                cell.number_format = '"$"#,##0.00'
    
    # Auto-adjust column widths
    for column in ws_main.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_main.column_dimensions[column_letter].width = adjusted_width
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Summary statistics
    total_amount = df['Amount'].sum()
    transaction_count = len(df)
    date_range = f"{df['Date'].min()} to {df['Date'].max()}" if not df.empty else "No data"
    
    # Category breakdown
    category_summary = df.groupby('Category')['Amount'].agg(['count', 'sum']).reset_index()
    
    # Account type breakdown
    account_summary = df.groupby('Account Type')['Amount'].agg(['count', 'sum']).reset_index()
    
    # Add summary data
    summary_data = [
        ["Export Summary", ""],
        ["Total Transactions", transaction_count],
        ["Total Amount", f"${total_amount:,.2f}"],
        ["Date Range", date_range],
        ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Category Breakdown", ""],
        ["Category", "Count", "Total Amount"]
    ]
    
    # Add category data
    for _, row in category_summary.iterrows():
        summary_data.append([row['Category'], row['count'], f"${row['sum']:,.2f}"])
    
    summary_data.extend([
        ["", ""],
        ["Account Type Breakdown", ""],
        ["Account Type", "Count", "Total Amount"]
    ])
    
    # Add account type data
    for _, row in account_summary.iterrows():
        summary_data.append([row['Account Type'], row['count'], f"${row['sum']:,.2f}"])
    
    # Write summary data
    for row_num, row_data in enumerate(summary_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_num, column=col_num, value=value)
            if row_num == 1 or (len(row_data) > 2 and value in ["Category Breakdown", "Account Type Breakdown"]):
                cell.font = Font(bold=True)
    
    # Auto-adjust summary column widths
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generate filename
    filename = f"lifetracker_transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Return as streaming response
    return StreamingResponse(
        io.BytesIO(excel_buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# PDF Processing Endpoint
@api_router.post("/transactions/pdf-import")
async def import_transactions_from_pdf(
    file: UploadFile = File(...),
    user_id: str = Depends(get_current_user_id)
):
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are supported")
    
    try:
        # Read PDF content
        content = await file.read()
        
        # Extract text from PDF
        text = extract_text_from_pdf(content)
        
        if not text.strip():
            raise HTTPException(status_code=400, detail="Could not extract text from PDF")
        
        # Parse transactions from text - pass the filename
        parsed_transactions = parse_transactions_from_text(text, user_id, file.filename)
        
        if not parsed_transactions:
            return {
                "message": "No transactions found in PDF", 
                "imported_count": 0,
                "extracted_text_preview": text[:500] + "..." if len(text) > 500 else text
            }
        
        # Check for duplicates and insert new transactions
        new_transactions = []
        duplicate_count = 0
        
        for trans_data in parsed_transactions:
            # Check if transaction already exists
            existing = await db.transactions.find_one({
                "user_id": user_id,
                "date": trans_data["date"],
                "description": trans_data["description"],
                "amount": trans_data["amount"]
            })
            
            if not existing:
                transaction_obj = Transaction(**trans_data)
                trans_dict = transaction_obj.dict()
                trans_dict['date'] = trans_dict['date'].isoformat() if hasattr(trans_dict['date'], 'isoformat') else trans_dict['date']
                trans_dict['created_at'] = trans_dict['created_at'].isoformat()
                new_transactions.append(trans_dict)
            else:
                duplicate_count += 1
        
        # Insert new transactions
        if new_transactions:
            await db.transactions.insert_many(new_transactions)
        
        return {
            "message": f"Successfully processed PDF: {file.filename}",
            "imported_count": len(new_transactions),
            "duplicate_count": duplicate_count,
            "total_found": len(parsed_transactions),
            "source_file": file.filename
        }
        
    except Exception as e:
        logging.error(f"PDF processing error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing PDF: {str(e)}")

# Enhanced Analytics (with user filtering)
@api_router.get("/analytics/monthly-report")
async def get_monthly_report(year: Optional[int] = None, user_id: str = Depends(get_current_user_id)):
    current_year = year or datetime.now().year
    
    # Get all transactions for the year
    start_date = f"{current_year}-01-01"
    end_date = f"{current_year}-12-31"
    
    transactions = await db.transactions.find({
        "user_id": user_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }).to_list(1000)
    
    # Group by month
    monthly_data = defaultdict(lambda: {
        "categories": defaultdict(float),
        "total_spent": 0,
        "transaction_count": 0
    })
    
    for transaction in transactions:
        trans_date = datetime.fromisoformat(transaction["date"]).date()
        month_key = f"{trans_date.year}-{trans_date.month:02d}"
        
        monthly_data[month_key]["categories"][transaction["category"]] += abs(transaction["amount"])
        monthly_data[month_key]["total_spent"] += abs(transaction["amount"])
        monthly_data[month_key]["transaction_count"] += 1
    
    # Convert to list format
    reports = []
    for month_key, data in monthly_data.items():
        year, month = month_key.split("-")
        reports.append({
            "month": month_key,
            "year": int(year),
            "categories": dict(data["categories"]),
            "total_spent": data["total_spent"],
            "transaction_count": data["transaction_count"]
        })
    
    return sorted(reports, key=lambda x: x["month"])

@api_router.get("/analytics/category-breakdown")
async def get_category_breakdown(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_id: str = Depends(get_current_user_id)
):
    filter_dict = {"user_id": user_id}
    if start_date:
        filter_dict["date"] = {"$gte": start_date}
    if end_date:
        if "date" in filter_dict:
            filter_dict["date"]["$lte"] = end_date
        else:
            filter_dict["date"] = {"$lte": end_date}
    
    transactions = await db.transactions.find(filter_dict).to_list(1000)
    
    category_data = defaultdict(lambda: {"amount": 0, "count": 0})
    total_spending = 0
    
    for transaction in transactions:
        amount = abs(transaction["amount"])
        category_data[transaction["category"]]["amount"] += amount
        category_data[transaction["category"]]["count"] += 1
        total_spending += amount
    
    # Calculate percentages and format response
    result = []
    for category, data in category_data.items():
        percentage = (data["amount"] / total_spending * 100) if total_spending > 0 else 0
        result.append({
            "category": category,
            "amount": data["amount"],
            "count": data["count"],
            "percentage": round(percentage, 2)
        })
    
    return sorted(result, key=lambda x: x["amount"], reverse=True)

# Enhanced bulk import with user support
@api_router.post("/transactions/bulk-import")
async def bulk_import_transactions(file: UploadFile = File(...), user_id: str = Depends(get_current_user_id)):
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are supported")
    
    try:
        content = await file.read()
        df = pd.read_csv(io.StringIO(content.decode('utf-8')))
        
        # Expected columns: date, description, category, amount
        required_columns = ['date', 'description', 'category', 'amount']
        if not all(col in df.columns for col in required_columns):
            raise HTTPException(
                status_code=400, 
                detail=f"CSV must contain columns: {', '.join(required_columns)}"
            )
        
        transactions = []
        for _, row in df.iterrows():
            # Convert date to string if it's a datetime object
            date_value = row['date']
            if hasattr(date_value, 'isoformat'):
                date_value = date_value.isoformat()
            elif hasattr(date_value, 'strftime'):
                date_value = date_value.strftime('%Y-%m-%d')
            
            transaction_data = {
                "date": str(date_value),
                "description": str(row['description']),
                "category": str(row['category']),
                "amount": float(row['amount']),
                "account_type": str(row.get('account_type', 'credit_card')),
                "user_id": user_id
            }
            transaction_obj = Transaction(**transaction_data)
            transaction_dict = transaction_obj.dict()
            # Ensure dates are strings for MongoDB
            if hasattr(transaction_dict['date'], 'isoformat'):
                transaction_dict['date'] = transaction_dict['date'].isoformat()
            if hasattr(transaction_dict['created_at'], 'isoformat'):
                transaction_dict['created_at'] = transaction_dict['created_at'].isoformat()
            transactions.append(transaction_dict)
        
        # Insert all transactions
        if transactions:
            await db.transactions.insert_many(transactions)
        
        return {"message": f"Successfully imported {len(transactions)} transactions"}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.get("/analytics/spending-trends")
async def get_spending_trends(months: int = 12, user_id: str = Depends(get_current_user_id)):
    # Get transactions from the last N months
    end_date = datetime.now().date()
    start_date = end_date.replace(month=end_date.month - months + 1 if end_date.month > months else 12 - (months - end_date.month - 1), 
                                  year=end_date.year if end_date.month > months else end_date.year - 1)
    
    transactions = await db.transactions.find({
        "user_id": user_id,
        "date": {"$gte": start_date.isoformat(), "$lte": end_date.isoformat()}
    }).to_list(1000)
    
    # Group by month for trend analysis
    monthly_trends = defaultdict(lambda: {"total": 0, "categories": defaultdict(float)})
    
    for transaction in transactions:
        trans_date = datetime.fromisoformat(transaction["date"]).date()
        month_key = f"{trans_date.year}-{trans_date.month:02d}"
        amount = abs(transaction["amount"])
        
        monthly_trends[month_key]["total"] += amount
        monthly_trends[month_key]["categories"][transaction["category"]] += amount
    
    return dict(monthly_trends)

# Account Type Analytics
@api_router.get("/analytics/account-type-breakdown")
async def get_account_type_breakdown(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_id: str = Depends(get_current_user_id)
):
    """Get spending breakdown by account type (debit vs credit)"""
    filter_dict = {"user_id": user_id}
    
    if start_date:
        filter_dict["date"] = {"$gte": start_date}
    if end_date:
        if "date" in filter_dict:
            filter_dict["date"]["$lte"] = end_date
        else:
            filter_dict["date"] = {"$lte": end_date}
    
    transactions = await db.transactions.find(filter_dict).to_list(10000)
    
    if not transactions:
        return {
            "debit": {"total": 0, "count": 0, "categories": {}},
            "credit": {"total": 0, "count": 0, "categories": {}}
        }
    
    account_breakdown = {
        "debit": {"total": 0, "count": 0, "categories": defaultdict(float)},
        "credit": {"total": 0, "count": 0, "categories": defaultdict(float)}
    }
    
    for transaction in transactions:
        amount = abs(transaction["amount"])
        account_type = "debit" if transaction.get("account_type") == "debit" else "credit"
        category = transaction["category"]
        
        account_breakdown[account_type]["total"] += amount
        account_breakdown[account_type]["count"] += 1
        account_breakdown[account_type]["categories"][category] += amount
    
    # Convert defaultdict to regular dict and calculate percentages
    result = {}
    total_spending = sum(acc["total"] for acc in account_breakdown.values())
    
    for account_type, data in account_breakdown.items():
        percentage = (data["total"] / total_spending * 100) if total_spending > 0 else 0
        result[account_type] = {
            "total": data["total"],
            "count": data["count"],
            "percentage": round(percentage, 2),
            "categories": dict(data["categories"])
        }
    
    return result

@api_router.get("/analytics/monthly-by-account-type")
async def get_monthly_breakdown_by_account_type(
    year: Optional[int] = None,
    user_id: str = Depends(get_current_user_id)
):
    """Get monthly spending breakdown by account type"""
    if year is None:
        year = datetime.now().year
    
    filter_dict = {
        "user_id": user_id,
        "date": {
            "$gte": f"{year}-01-01",
            "$lte": f"{year}-12-31"
        }
    }
    
    transactions = await db.transactions.find(filter_dict).to_list(10000)
    
    # Group by month and account type
    monthly_data = defaultdict(lambda: {
        "debit": {"total": 0, "count": 0},
        "credit": {"total": 0, "count": 0}
    })
    
    for transaction in transactions:
        trans_date = datetime.fromisoformat(transaction["date"]).date()
        month_key = f"{trans_date.year}-{trans_date.month:02d}"
        amount = abs(transaction["amount"])
        account_type = "debit" if transaction.get("account_type") == "debit" else "credit"
        
        monthly_data[month_key][account_type]["total"] += amount
        monthly_data[month_key][account_type]["count"] += 1
    
    # Convert to list format
    reports = []
    for month_key, data in monthly_data.items():
        year_val, month_val = month_key.split("-")
        reports.append({
            "month": month_key,
            "year": int(year_val),
            "debit": data["debit"],
            "credit": data["credit"],
            "total": data["debit"]["total"] + data["credit"]["total"]
        })
    
    return sorted(reports, key=lambda x: x["month"])

@api_router.get("/analytics/source-breakdown")
async def get_source_breakdown(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_id: str = Depends(get_current_user_id)
):
    """Get spending breakdown by transaction source (e.g., Jane's Debit, John's Credit)"""
    filter_dict = {"user_id": user_id}
    
    if start_date:
        filter_dict["date"] = {"$gte": start_date}
    if end_date:
        if "date" in filter_dict:
            filter_dict["date"]["$lte"] = end_date
        else:
            filter_dict["date"] = {"$lte": end_date}
    
    transactions = await db.transactions.find(filter_dict).to_list(10000)
    
    if not transactions:
        return []
    
    source_breakdown = defaultdict(lambda: {"total": 0, "count": 0, "account_type": ""})
    
    for transaction in transactions:
        amount = abs(transaction["amount"])
        source = transaction.get("pdf_source", "Manual")
        account_type = transaction.get("account_type", "unknown")
        
        source_breakdown[source]["total"] += amount
        source_breakdown[source]["count"] += 1
        source_breakdown[source]["account_type"] = account_type
    
    # Convert to list and calculate percentages
    total_spending = sum(data["total"] for data in source_breakdown.values())
    
    result = []
    for source, data in source_breakdown.items():
        percentage = (data["total"] / total_spending * 100) if total_spending > 0 else 0
        result.append({
            "source": source,
            "total": data["total"],
            "count": data["count"],
            "percentage": round(percentage, 2),
            "account_type": data["account_type"]
        })
    
    return sorted(result, key=lambda x: x["total"], reverse=True)

# Authentication endpoints
auth_router = APIRouter(prefix="/auth", tags=["authentication"])

@auth_router.post("/register", response_model=User)
async def register_user(user_data: UserCreate):
    # Check if user already exists
    existing_user = await get_user_by_email(user_data.email)
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email already registered"
        )
    
    # Create username from email if not provided
    username = user_data.username or user_data.email.split("@")[0]
    
    # Check if username is taken
    existing_username = await db.users.find_one({"username": username})
    if existing_username:
        # Add random suffix if username exists
        username = f"{username}_{str(uuid.uuid4())[:8]}"
    
    # Hash password
    hashed_password = get_password_hash(user_data.password)
    
    # Create user document
    user_doc = {
        "id": str(uuid.uuid4()),
        "email": user_data.email,
        "username": username,
        "full_name": user_data.full_name,
        "password_hash": hashed_password,
        "role": "user",
        "household_id": None,  # Will be set when user joins/creates household
        "created_at": datetime.utcnow(),
        "last_login": None
    }
    
    await db.users.insert_one(user_doc)
    
    # Return user without password
    return User(**{k: v for k, v in user_doc.items() if k != "password_hash"})

@auth_router.post("/login", response_model=Token)
async def login_user(form_data: OAuth2PasswordRequestForm = Depends()):
    user = await authenticate_user(form_data.username, form_data.password)  # username field contains email
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Update last login
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"last_login": datetime.utcnow()}}
    )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["email"]}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

@auth_router.get("/me", response_model=User)
async def get_current_user_info(current_user: dict = Depends(get_current_user)):
    return User(**{k: v for k, v in current_user.items() if k != "password_hash"})

@auth_router.post("/household", response_model=Household)
async def create_household(household_data: HouseholdCreate, current_user: dict = Depends(get_current_user)):
    household_doc = {
        "id": str(uuid.uuid4()),
        "name": household_data.name,
        "created_by": current_user["id"],
        "members": [current_user["id"]],
        "created_at": datetime.utcnow()
    }
    
    await db.households.insert_one(household_doc)
    
    # Update user's household_id
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"household_id": household_doc["id"]}}
    )
    
    return Household(**household_doc)

@auth_router.get("/household", response_model=Optional[Household])
async def get_user_household(current_user: dict = Depends(get_current_user)):
    if not current_user.get("household_id"):
        return None
    
    household = await db.households.find_one({"id": current_user["household_id"]})
    if household:
        return Household(**household)
    return None

@auth_router.get("/household/members", response_model=List[User])
async def get_household_members(current_user: dict = Depends(get_current_user)):
    if not current_user.get("household_id"):
        return []
    
    # Get all users in the same household
    users = await db.users.find({"household_id": current_user["household_id"]}).to_list(100)
    
    # Return users without password hashes
    return [User(**{k: v for k, v in user.items() if k != "password_hash"}) for user in users]

@auth_router.post("/household/invite")
async def invite_household_member(
    invitation: dict,
    current_user: dict = Depends(get_current_user)
):
    if not current_user.get("household_id"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="User is not part of a household"
        )
    
    email = invitation.get("email")
    role = invitation.get("role", "user")
    
    # Check if user already exists
    existing_user = await get_user_by_email(email)
    if existing_user:
        if existing_user.get("household_id"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User is already part of a household"
            )
        
        # Add existing user to household
        await db.users.update_one(
            {"id": existing_user["id"]},
            {"$set": {"household_id": current_user["household_id"], "role": role}}
        )
        
        # Add user to household members list
        await db.households.update_one(
            {"id": current_user["household_id"]},
            {"$addToSet": {"members": existing_user["id"]}}
        )
        
        return {"message": "User added to household successfully", "user_id": existing_user["id"]}
    else:
        # Create invitation record (for future implementation)
        invitation_doc = {
            "id": str(uuid.uuid4()),
            "email": email,
            "household_id": current_user["household_id"],
            "invited_by": current_user["id"],
            "role": role,
            "status": "pending",
            "created_at": datetime.utcnow()
        }
        
        await db.invitations.insert_one(invitation_doc)
        
        return {"message": "Invitation sent", "invitation_id": invitation_doc["id"]}

# Password Reset Endpoints
@auth_router.post("/forgot-password")
async def forgot_password(request: PasswordResetRequest):
    user = await get_user_by_email(request.email)
    if not user:
        # Don't reveal if email exists for security
        return {"message": "If the email exists, a reset code has been sent"}
    
    # Generate reset code
    reset_code = generate_reset_code()
    
    # Store reset code in database with expiration
    reset_doc = {
        "email": request.email,
        "reset_code": reset_code,
        "created_at": datetime.utcnow(),
        "expires_at": datetime.utcnow() + timedelta(minutes=15)  # 15 minute expiry
    }
    
    # Remove any existing reset codes for this email
    await db.password_resets.delete_many({"email": request.email})
    await db.password_resets.insert_one(reset_doc)
    
    # Send email
    subject = "LifeTracker - Password Reset Code"
    body = f"""
    <html>
    <body>
        <h2>Password Reset Request</h2>
        <p>You requested a password reset for your LifeTracker account.</p>
        <p>Your reset code is: <strong style="font-size: 18px; color: #3B82F6;">{reset_code}</strong></p>
        <p>This code will expire in 15 minutes.</p>
        <p>If you didn't request this reset, please ignore this email.</p>
        <br>
        <p>Best regards,<br>LifeTracker Team</p>
    </body>
    </html>
    """
    
    await send_email(request.email, subject, body)
    return {"message": "If the email exists, a reset code has been sent"}

@auth_router.post("/reset-password")
async def reset_password(request: PasswordResetConfirm):
    # Find valid reset code
    reset_record = await db.password_resets.find_one({
        "email": request.email,
        "reset_code": request.reset_code,
        "expires_at": {"$gt": datetime.utcnow()}
    })
    
    if not reset_record:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid or expired reset code"
        )
    
    # Update user password
    hashed_password = get_password_hash(request.new_password)
    await db.users.update_one(
        {"email": request.email},
        {"$set": {"password_hash": hashed_password}}
    )
    
    # Remove used reset code
    await db.password_resets.delete_many({"email": request.email})
    
    return {"message": "Password successfully reset"}

# User Profile Management Endpoints
@auth_router.put("/profile", response_model=User)
async def update_user_profile(
    profile_update: UserProfileUpdate, 
    current_user: dict = Depends(get_current_user)
):
    update_data = {}
    
    if profile_update.full_name is not None:
        update_data["full_name"] = profile_update.full_name
    
    if profile_update.username is not None:
        # Check if username is taken by another user
        existing_user = await db.users.find_one({
            "username": profile_update.username,
            "id": {"$ne": current_user["id"]}
        })
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Username already taken"
            )
        update_data["username"] = profile_update.username
    
    if update_data:
        await db.users.update_one(
            {"id": current_user["id"]},
            {"$set": update_data}
        )
        
        # Get updated user
        updated_user = await db.users.find_one({"id": current_user["id"]})
        return User(**{k: v for k, v in updated_user.items() if k != "password_hash"})
    
    return User(**{k: v for k, v in current_user.items() if k != "password_hash"})

@auth_router.post("/change-password")
async def change_password(
    password_request: PasswordChangeRequest,
    current_user: dict = Depends(get_current_user)
):
    # Verify current password
    if not verify_password(password_request.current_password, current_user["password_hash"]):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Current password is incorrect"
        )
    
    # Update password
    hashed_password = get_password_hash(password_request.new_password)
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"password_hash": hashed_password}}
    )
    
    return {"message": "Password successfully changed"}

# Google OAuth Endpoints
@auth_router.get("/google/login")
async def google_login(request: Request):
    # Create OAuth config
    oauth = OAuth()
    google = oauth.register(
        name='google',
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        server_metadata_url='https://accounts.google.com/.well-known/openid_configuration',
        client_kwargs={
            'scope': 'openid email profile'
        }
    )
    
    # Build redirect URI
    redirect_uri = f"{str(request.base_url).rstrip('/')}/auth/google/callback"
    return await google.authorize_redirect(request, redirect_uri)

@auth_router.get("/google/callback")
async def google_callback(request: Request):
    try:
        # Create OAuth config
        oauth = OAuth()
        google = oauth.register(
            name='google',
            client_id=GOOGLE_CLIENT_ID,
            client_secret=GOOGLE_CLIENT_SECRET,
            server_metadata_url='https://accounts.google.com/.well-known/openid_configuration',
            client_kwargs={
                'scope': 'openid email profile'
            }
        )
        
        # Get token and user info
        token = await google.authorize_access_token(request)
        user_info = token.get('userinfo')
        
        if not user_info or not user_info.get('email'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Failed to get user information from Google"
            )
        
        email = user_info['email']
        full_name = user_info.get('name', '')
        username = email.split('@')[0]
        
        # Check if user exists
        existing_user = await get_user_by_email(email)
        
        if existing_user:
            # Update last login
            await db.users.update_one(
                {"id": existing_user["id"]},
                {"$set": {"last_login": datetime.utcnow()}}
            )
            user_data = existing_user
        else:
            # Create new user
            # Check if username is taken
            existing_username = await db.users.find_one({"username": username})
            if existing_username:
                username = f"{username}_{str(uuid.uuid4())[:8]}"
            
            user_doc = {
                "id": str(uuid.uuid4()),
                "email": email,
                "username": username,
                "full_name": full_name,
                "password_hash": get_password_hash(str(uuid.uuid4())),  # Random password for OAuth users
                "role": "user",
                "household_id": None,
                "created_at": datetime.utcnow(),
                "last_login": datetime.utcnow()
            }
            
            await db.users.insert_one(user_doc)
            user_data = user_doc
        
        # Create JWT token
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        access_token = create_access_token(
            data={"sub": user_data["email"]}, expires_delta=access_token_expires
        )
        
        # Return token and redirect to frontend with token in query params
        frontend_url = os.environ.get('REACT_APP_BACKEND_URL', 'http://localhost:3000').replace('/api', '').replace(':8001', ':3000')
        redirect_url = f"{frontend_url}?token={access_token}"
        
        return {"redirect_url": redirect_url, "access_token": access_token, "token_type": "bearer"}
        
    except Exception as e:
        logging.error(f"Google OAuth callback error: {e}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="OAuth authentication failed"
        )

# Legacy endpoint for backward compatibility (will be removed later)
async def get_current_user_id_legacy():
    """Legacy function for backward compatibility during migration"""
    return "default_user"

# Include both routers in the main app
app.include_router(api_router)
app.include_router(auth_router, prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[
        "http://localhost:3000",  # Local development
        "https://*.vercel.app",   # Vercel deployments
        "https://*.netlify.app",  # Netlify deployments
        "*"  # Allow all for now (change in production)
    ],
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()